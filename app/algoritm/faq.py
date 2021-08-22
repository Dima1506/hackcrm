import string
import scipy.spatial.distance as ds
import numpy as np
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from navec import Navec
import random

#sentence_avg_vector = avg_sentence_vector(senp, model=model, num_features=300,index2word_set=model.vocab) #senp это строка, 

def sentence_avg_vector(words, num_features = 300):
    words = words.translate(str.maketrans('', '', string.punctuation))
    words = words.lower().split(' ')
    #function to average all words vectors in a given paragraph
    featureVec = np.zeros((num_features,), dtype="float32")
    nwords = 0
    for word in words:
        try:
            featureVec = np.add(featureVec, model[word])
            nwords = nwords+1
        except:
            print('error')

    if nwords>0:
        featureVec = np.divide(featureVec, nwords)
    return featureVec

def sentence_middle_avg(words_mas, vector, num_features = 300):
    featureVec = np.zeros((num_features,), dtype="float32")
    for sentence_id in words_mas:
        featureVec = np.add(featureVec, vector[sentence_id])
    featureVec = np.divide(featureVec, len(words_mas))
    return featureVec

def make_all_vectorwords(words, types):
    '''for s in words:
        t = sentence_avg_vector(s.lower())
        if 0 in t:
            print(s, t)'''
    vector_arr = np.array([sentence_avg_vector(s.lower()) for s in words])
    print('sd')
    np.save(types, vector_arr)
    #print(vector_arr)


def sentence_faq(sentence):
    vector = np.load('.npy')
    min_senstence = 1.0
    state_min = 0
    sentence = sentence.lower()
    sentence_avg_vector = avg_sentence_vector(sentence, model=model, num_features=300,index2word_set=model.vocab)
    #print(len(vector))
    for sen in range(len(vector)):
        sen1_sen2_similarity = ds.cosine(sentence_avg_vector, vector[sen])#cosine_similarity(sentence_1_avg_vector,sentence_2_avg_vector)
        #print(sen1_sen2_similarity)
        if min_senstence > sen1_sen2_similarity:
            min_senstence = sen1_sen2_similarity
            state_min = sen
    return(question[state_min])



def parse_xlsx(sentence_json):
    wb = load_workbook('text.xlsx')
    ws = wb.active
    #print(ws.max_column)
    column = ws.max_column
    mass = []
    types=''
    print(get_column_letter(column))
    column = ws[get_column_letter(column)]
    id_sentence = -1
    print(len(column))
    for i in range(len(column)):
        if i == 0:
            types=column[i].value
        else:
            strp=''
            pred_element=''
            ka=0
            for char in column[i].value:
                if char=='.' or char=='?' or char=='!':
                    strp+=char
                    mass.append(strp)
                    strp=''
                    id_sentence+=1
                    if ka == 0:
                        sentence_json['zero_id_sentence'].append(id_sentence)
                        ka = 1
                elif (pred_element=='.' or pred_element=='?' or pred_element=='!') and (char==' '):
                    continue
                else:
                    strp+=char
                pred_element = char
            #print(id_sentence, column[i].value)
            sentence_json['end_id_sentence'].append(id_sentence)
            #print(mass)
            #mass = mass + re.split("r'[ .]+", column[i].value)
            #mass.append(.split())
        sentence_json['text']=mass
    for i in range(len(mass)):
        if mass[i][0:1]=='':
            print(mass[i],mass[i+1])
    #print(mass)
    #print(types)
    make_all_vectorwords(mass,types)
    avg_class(sentence_json)
    return sentence_json

def avg_class(sentence_json):
    class_number = 0
    exclude_mas = []
    for i in range(len(sentence_json['text'])):
        #print('inter1 '+str(i))
        ka = 0
        exclude_mas.append(i)
        #print(exclude_mas)
        #print(exclude_mas)
        #print('sen  '+str(sentence_json['sentence_avg']))
        for j in range(len(sentence_json['text'])):
            if j not in exclude_mas:
                vector = np.load('Развлекательный.npy')
                sen1_sen2_similarity = ds.cosine(vector[i], vector[j])
                if sen1_sen2_similarity<0.14:
                    #print('uk')
                    #print(i,j)
                    #print('ul')
                    #print('test'+str(sentence_json['sentence_avg']))
                    #print('test2'+str(sentence_json['class_avg']))
                    #print(str(i))
                    #print('test3'+str(sentence_json['class_avg'][str(i)]))
                    if str(i) in sentence_json['class_avg']:
                        massiv_class = sentence_json['sentence_avg'][sentence_json['class_avg'][str(i)]]
                        #print('sss '+str(massiv_class))
                        sen1_sen_Avgclass_similarity = ds.cosine(vector[i], sentence_middle_avg(massiv_class, vector)) 
                        #print('avg_sred'+str(sen1_sen_Avgclass_similarity))
                        if sen1_sen_Avgclass_similarity < 0.1:
                            if str(j) not in sentence_json['sentence_avg']:
                                mas_vremen = sentence_json['sentence_avg'][sentence_json['class_avg'][str(i)]]
                                #print(mas_vremen)
                                mas_vremen.append(j)
                                #print(j)
                                sentence_json['sentence_avg'][sentence_json['class_avg'][str(i)]] = mas_vremen
                                sentence_json['class_avg'][str(j)]=str(sentence_json['class_avg'][str(i)])
                                print('1:'+sentence_json['text'][i]+'\n'+'2:'+sentence_json['text'][j])
                                print('1bug:'+sentence_json['text'][i+1]+'\n'+'2bug:'+sentence_json['text'][j+1])
                    elif str(j) in sentence_json['class_avg']:
                        massiv_class = sentence_json['sentence_avg'][sentence_json['class_avg'][str(j)]]
                        #print('sss2 '+str(massiv_class))
                        sen1_sen_Avgclass_similarity = ds.cosine(vector[j], sentence_middle_avg(massiv_class, vector)) 
                        #print('avg_sred2 '+str(sen1_sen_Avgclass_similarity))
                        if sen1_sen_Avgclass_similarity < 0.1:
                            if str(i) not in sentence_json['sentence_avg']:
                                #print(str(j))
                                mas_vremen = sentence_json['sentence_avg'][sentence_json['class_avg'][str(j)]]
                                #print(mas_vremen)
                                mas_vremen.append(i)
                                #print(i)

                                sentence_json['sentence_avg'][sentence_json['class_avg'][str(j)]] = mas_vremen
                                sentence_json['class_avg'][str(i)]=str(sentence_json['class_avg'][str(j)])
                                print('1:'+sentence_json['text'][i]+'\n'+'2:'+sentence_json['text'][j])
                                print('1bug:'+sentence_json['text'][i+1]+'\n'+'2bug:'+sentence_json['text'][j+1])
                    else:
                        ka = 1
                        sentence_json['class_avg'][str(i)]=str(class_number)
                        sentence_json['class_avg'][str(j)]=str(class_number)
                        sentence_json['sentence_avg'][str(class_number)]=[i,j]
                        class_number+=1
                        #exclude_mas.append(j)
                        print('1:'+sentence_json['text'][i]+'\n'+'2:'+sentence_json['text'][j])
                        print('1bug:'+sentence_json['text'][i+1]+'\n'+'2bug:'+sentence_json['text'][j+1])
                        #print(sen1_sen2_similarity)
                        #print(i,j)
        if ka == 0:
            sentence_json['class_avg'][str(i)]=str(class_number)
            sentence_json['sentence_avg'][str(class_number)]=[i]
            class_number+=1
    return sentence_json

        
def markov_form_service(sentence_json):
    #sentence_json['sentence_avg']['inter']= sentence_json['zero_id_sentence']
    #sentence_json['sentence_avg']['end']= sentence_json['zero_id_sentence']
    sentence_json['sentence_markov']['start'] = {'next_sent':[]}
    for i in range(len(sentence_json['text'])):
        if i in sentence_json['zero_id_sentence']:
            sentence_json['sentence_markov'][sentence_json['class_avg'][str(i)]] = {
            'sent':sentence_json['sentence_avg'][sentence_json['class_avg'][str(i)]],
                'next_sent':[]
            }
            for j in sentence_json['sentence_markov'][sentence_json['class_avg'][str(i)]]['sent']:
                sentence_json['sentence_markov'][sentence_json['class_avg'][str(i)]]['next_sent'].append(sentence_json['class_avg'][str(j+1)])
            sentence_json['sentence_markov']['start']['next_sent'].append(sentence_json['class_avg'][str(i)])
        elif i in sentence_json['end_id_sentence']:
            if sentence_json['class_avg'][str(i)] not in sentence_json['sentence_markov']:
                sentence_json['sentence_markov'][sentence_json['class_avg'][str(i)]] = {
                'sent':sentence_json['sentence_avg'][sentence_json['class_avg'][str(i)]],
                    'next_sent':['end']
                }
            else:
                #sentence_json['sentence_markov'][sentence_json['class_avg'][str(i)]]['sent'].append(str(i))
                sentence_json['sentence_markov'][sentence_json['class_avg'][str(i)]]['next_sent'].append('end')
        else:
            sentence_json['sentence_markov'][sentence_json['class_avg'][str(i)]] = {
            'sent':sentence_json['sentence_avg'][sentence_json['class_avg'][str(i)]],
                'next_sent':[]
            }
            for j in sentence_json['sentence_markov'][sentence_json['class_avg'][str(i)]]['sent']:
                sentence_json['sentence_markov'][sentence_json['class_avg'][str(i)]]['next_sent'].append(sentence_json['class_avg'][str(j+1)])
            #sentence_json['sentence_markov']['start']['next_sent'].append(sentence_json['class_avg'][str(i)])
    print(sentence_json['sentence_markov']['0'])
    return sentence_json

def markov_cep():
    '''path = 'navec_hudlit_v1_12B_500K_300d_100q.tar'
    model = Navec.load(path)'''
    f = open("algoritm/data.txt", "r")
    sentence_json = eval(f.read())
    f.close()
    class_s = random.choice(sentence_json['sentence_markov']['start']['next_sent'])
    text = ''
    while True:
        num_sentences = random.choice(sentence_json['sentence_markov'][class_s]['sent'])
        class_new = random.choice(sentence_json['sentence_markov'][class_s]['next_sent'])
        text += sentence_json['text'][num_sentences]
        class_s = class_new
        if class_s == 'end':
            break
    return text


'''
#make_all_vectorwords()
#print(sentence_faq('хочу отдохнуть'))
sentence_json={'sentence_markov':{},'sentences':[], 'sentence_avg':{}, 'zero_id_sentence':[],'class_avg':{},'end_id_sentence':[]}
sentence_json = parse_xlsx(sentence_json)
f = open("data.txt", "w")
f.write(str(sentence_json))
f.close()
f = open("data.txt", "r")
sentence_json = eval(f.read())
f.close()
sentence_json = markov_form_service(sentence_json)

print(markov_cep(sentence_json))
#print(sentence_json)'''

